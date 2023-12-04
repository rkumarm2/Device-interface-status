#!/usr/bin/env python
# coding: utf-8

import tk
import tkinter as tk
from tkinter import *
from tkinter import ttk

import aide
import openpyxl
import xlrd as xl
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfiles

import xlsxwriter
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


aide.submit_statistics(
        pid = "933120",
        tool_id="656d70036887b1184c355cb3",
        metadata={
            "potential_savings": 4,  # Hours
            "report_savings": True,
        },
    )


def open_file():
    browse_text.set("loading...")
    
    file = askopenfiles(parent=root, mode='rb', title="`Choose a file")

    for i in file:
        i=i.name
        t= i.split("/")
        if(t[-1][-4:-1] == "xls"):
            ssfile = i
            ssfilename = t[-1]
        else:
            tfile = i
            tfilename = t[-1]
    wb = load_workbook(ssfile)
    ws = wb["LLD"]
    rc = ws.max_row
    cc = ws.max_column
    print("Text file : ", tfilename)
    print("Spreadsheet: ", ssfilename)
    oldiflist , newiflist , chkdupl , duplist = [], [] , [] , []
    
    for i in range(2,rc+1):
        acell= "A" + str(i)
        bcell= "B" + str(i)
        ac = ws[acell].value 
        bc = ws[bcell].value
        oldiflist.append( ws[acell].value )
        newiflist.append( ws[bcell].value )
        
    for i in range(rc-1) :
        chkdupl.append(  newiflist.count(newiflist[i]) )
  
    wb.close()
    f=open(tfile) 
    txtlist = [] 
    counter=0
    for i in f.readlines():
        l=i.split(" ")
        flag=0 
        length = len(l)
        for word in range(length):
            if l[word][:7] == "Hundred":
                existing = "Hu"+ l[word][11:]
                if existing in oldiflist:
                    new_word_ind = oldiflist.index( existing)
                    newword = newiflist[ new_word_ind ]
                    rep_count = chkdupl[ new_word_ind ]
                    
                    if rep_count == 1 :
                        l[word] = newword 
                    else:
                        l[word] = l[word][:11] + "_Recheck_" + newword + "_old_" + l[word]
                        l= [existing] + l
                        duplist.append(l)
                        txtlist.append(["\n!Duplicate was here, shifted to the bottom page !\n"])
                        flag=1
                else:
                    l[word] = l[word][:11]+"_Existing_"+l[word][11:]
        if flag == 0 :
            txtlist.append(l)
    f.close()
    
    file = open("Devicename_Targetconfig.txt","w")
    for i in txtlist:
        file.write(" ".join(i)) 
    
    duplist.sort()
    if len( duplist) > 0 : 
        file.write("\n\n!!! Duplicates are below !!! \n\n")
        print("\nThere are duplicates in the file. \n")
        x = duplist[0][0]  
    for i in duplist:
        if i[0] != x:
            file.write("\n\n! \n")
            x=i[0]
        file.write(" ".join(i))
    file.close()
    
    browse_text.set("Completed")
    createdText = tk.Label(root, text ="File created successfully",font="Raleway")
    createdText.grid(columnspan=3 , column=1 , row = 5)




root=tk.Tk()

canvas = tk.Canvas(root,width=660,height = 620)
canvas.grid(columnspan=3,rowspan=5)
logo = Image.open('switchNrouter.jpg')
logo = ImageTk.PhotoImage(logo)
logo_label=tk.Label(image=logo)
logo_label.image =logo
logo_label.grid(column=1,row=0)

instructions = tk.Label(root, text ="Select the spreadsheet and txt files",font="Raleway")
instructions.grid(columnspan=3,column=0,row=2)

browse_text = tk.StringVar()
browse_btn = tk.Button(root, textvariable=browse_text, command=lambda:open_file(), font ="Raleway", bg="green", fg="white",height=2,width=12)
browse_text.set("Browse")
browse_btn.grid(columnspan=3,column=1,row=3)

root.mainloop()
